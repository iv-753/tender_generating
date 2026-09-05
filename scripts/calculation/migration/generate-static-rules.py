"""One-time migration helper: transcribe workbook rows into static ESM rule tables."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


HERE = Path(__file__).resolve()
WORKBOOK = next(
    (parent / "动态成本分析模型.xlsx" for parent in HERE.parents
     if (parent / "动态成本分析模型.xlsx").is_file()),
    HERE.parents[4] / "动态成本分析模型.xlsx",
)
OUTPUT = HERE.parents[1] / "rules"
PARAMETER_MAP = HERE.with_name("full-model-parameter-map.json")
GRADES = ("A", "B", "C", "D")
PRICE_SHEET_NAME = "分级单价保洁、绿化 "
MISSING_SHEETS = {
    "四害消杀": (5, 11, "pest-control"),
    "工程委外": (5, 99, "engineering-outsourced"),
    "工程常规": (5, 232, "engineering-routine"),
}
GRADE_REFERENCE = re.compile(r"'分级单价保洁、绿化 '!([A-Z]+[0-9]+)")

PARAMETER_LABELS = {
    "pest.treatmentArea": "四害消杀面积",
    "basement.parkingArea": "地下停车区面积",
    "basement.fireShutterCount": "地下停车区防火卷帘数量",
    "basement.parkingSurveillanceCount": "地下停车区监控设施数量",
    "basement.vehicleEntranceEquipmentCount": "地下车道出入口设备数量",
    "basement.generatorRoomCount": "地下发电机房数量",
    "basement.transformerRoomCount": "地下专变电房数量",
    "basement.telecomRoomCount": "地下电信机房数量",
    "basement.domesticWaterPumpRoomCount": "地下生活给水泵房数量",
    "basement.poolPumpRoomCount": "地下泳池泵房数量",
    "basement.firePumpRoomCount": "地下消防水泵房数量",
    "basement.wetAlarmValveRoomCount": "地下湿式报警阀间数量",
    "basement.powerDistributionRoomCount": "地下配电间数量",
    "basement.chargingMeterRoomCount": "地下充电桩电表间数量",
    "basement.pressurizationFanRoomCount": "地下加压风机房数量",
    "basement.smokeExhaustFanRoomCount": "地下排烟风机房数量",
    "basement.supplyAirFanRoomCount": "地下送风风机房数量",
    "basement.sewageTreatmentRoomCount": "地下污水处理机房数量",
    "basement.rainwaterStorageRoomCount": "地下雨水调蓄机房数量",
    "basement.lobbyAccessControlCount": "地下大堂门禁设备数量",
    "basement.lobbyEquipmentCount": "地下大堂设备数量",
    "basement.lobbyArea": "地下大堂面积",
    "basement.drainageEquipmentCount": "地下排水排污设备数量",
    "basement.evacuationStairArea": "地下疏散楼梯面积",
    "basement.vehicleEntranceArea": "地下车道出入口面积",
    "building.groundFloorLobbyAccessControlCount": "首层大堂门禁设备数量",
    "building.groundFloorLobbyArea": "首层大堂面积",
    "building.stiltFloorArea": "首层架空层面积",
    "building.clubhouseCount": "首层泛会所数量",
    "building.securityControlRoomCount": "安防控制室数量",
    "building.groundFloorElectricalRoomCount": "首层电房数量",
    "building.detachedElectricalRoomCount": "独立配套用房电房数量",
    "building.standardFloorArea": "标准层公区面积",
    "building.roofArea": "屋面层面积",
    "building.buildingCount": "楼栋数量",
    "building.evacuationStairArea": "楼栋疏散楼梯面积",
    "building.elevatorCount": "电梯数量",
    "building.refugeFloorArea": "避难层面积",
    "building.refugeWetAlarmRoomCount": "避难层湿式报警间数量",
    "building.refugePressurizationFanRoomCount": "避难层加压风机房数量",
    "building.commercialCorridorArea": "商业公区走廊面积",
    "building.commercialEvacuationStairArea": "商业公区疏散楼梯面积",
    "building.commercialEvacuationStairMaintenanceArea": "商业公区疏散楼梯保养面积",
    "building.shaftCount": "楼栋管井数量",
    "building.facadeBaseArea": "外立面基座层面积",
    "building.facadeStandardArea": "外立面标准层面积",
    "building.facadeRoofParapetArea": "外立面屋顶女儿墙面积",
    "building.commercialElectricalShaftCount": "商业公区电井数量",
    "building.commercialWaterShaftCount": "商业公区水井数量",
    "building.commercialRoofArea": "商业公区屋面面积",
    "building.commercialFacadeArea": "商业外立面面积",
    "building.garbageTransferStationCount": "垃圾中转站数量",
    "grounds.gatedEntrancePedestrianGateCount": "有门楼出入口人行闸机数量",
    "grounds.gatedEntranceVehicleGateCount": "有门楼出入口车行闸机数量",
    "grounds.ungatedEntrancePedestrianGateCount": "无门楼出入口人行闸机数量",
    "grounds.ungatedEntranceVehicleGateCount": "无门楼出入口车行闸机数量",
    "grounds.gatedEntranceAccessGateCount": "有门楼出入口通行门数量",
    "grounds.ungatedEntranceAccessGateCount": "无门楼出入口通行门数量",
    "grounds.securityAudioLineLength": "园区安防及背景音箱线路长度",
    "grounds.fireWaterPointCount": "园区消防取水设施数量",
    "grounds.entranceGuardhouseArea": "出入口门岗及廊架面积",
    "grounds.entrancePlazaArea": "出入口及广场面积",
    "grounds.entranceWaterFeatureCount": "出入口及广场水景数量",
    "grounds.landscapeWaterFeatureCount": "园林水景数量",
    "grounds.entranceArtworkCount": "出入口及广场艺术雕塑数量",
    "grounds.landscapeArtworkCount": "园林设施艺术雕塑数量",
    "grounds.entranceBicycleShelterCount": "出入口及广场非机动车棚数量",
    "grounds.landscapeBicycleShelterCount": "园林设施非机动车棚数量",
    "grounds.zoneArea": "园林分区面积",
    "grounds.poolPergolaArea": "泳池廊架及地面面积",
    "grounds.poolEquipmentCount": "泳池设备数量",
    "grounds.poolFacilityRoomCount": "泳池功能房数量",
    "grounds.childrenActivityArea": "儿童活动场地面积",
    "grounds.activityFacilityArea": "儿童及健身活动设施面积",
    "grounds.fitnessActivityArea": "全民健身活动场地面积",
    "grounds.badmintonCourtCount": "羽毛球场数量",
    "grounds.basketballCourtCount": "篮球场数量",
    "grounds.tennisCourtCount": "网球场数量",
    "grounds.footballFieldCount": "足球场数量",
    "grounds.vehicleParkingArea": "园区机动车停车场面积",
    "grounds.pergolaPlatformArea": "园区廊架平台面积",
    "grounds.waterOutletCount": "园区取水栓及排水口数量",
    "grounds.wasteCollectionPointCount": "园区垃圾收集点数量",
    "grounds.signageCount": "园区标识标牌及宣传栏数量",
    "grounds.roadArea": "园区道路面积",
    "grounds.wallLength": "园区围墙长度",
    "grounds.tankCount": "园区池体数量",
    "grounds.drainagePipelineLength": "园区排水主管长度",
    "grounds.drainageWellCount": "园区排水井数量",
    "grounds.distributionBoxCount": "园林配电箱数量",
}


def js(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def write_module(name: str, export_name: str, rows: list[dict]) -> None:
    content = (
        "// Generated once from the internal workbook; production never reads that workbook.\n"
        f"export const {export_name} = Object.freeze({js(rows)}.map(Object.freeze));\n"
    )
    (OUTPUT / name).write_text(content, encoding="utf-8", newline="\n")


def multiplier(formula: str) -> float:
    match = re.search(r"\*([0-9.]+)$", formula)
    if not match:
        raise ValueError(f"Cannot find multiplier: {formula}")
    return float(match.group(1))


def ratio(formula: str) -> float:
    match = re.search(r"\*([0-9.]+)$", formula)
    return float(match.group(1)) if match else 0.0


def grade_values(sheet, row: int, columns: tuple[int, int, int, int]) -> dict:
    return dict(zip(GRADES, (sheet.cell(row, column).value for column in columns)))


def service_rules(sheet) -> list[dict]:
    demands = {
        5: {"type": "linear", "terms": {"occupiedHouseholds": 0.5}},
        6: {"type": "linear", "terms": {"deliveredHouseholds": 0.02, "receivedHouseholds": -0.01, "occupiedHouseholds": -0.005}},
        7: {"type": "linear", "terms": {"occupiedHouseholds": 1}},
        8: {"type": "constant", "value": 0},
        9: {"type": "linear", "terms": {"deliveredHouseholds": 0.5, "occupiedHouseholds": -0.4}},
        10: {"type": "linear", "terms": {"occupiedHouseholds": 3}},
        11: {"type": "linear", "terms": {"deliveredHouseholds": 1, "receivedHouseholds": 1, "occupiedHouseholds": 4}},
        12: {"type": "grade-linear", "terms": {"receivedHouseholds": 0.05, "occupiedHouseholds": 0.15}, "gradeFactors": {"A": 1.5, "B": 1.2, "C": 1, "D": 0.8}},
        13: {"type": "grade-linear", "terms": {"occupiedHouseholds": 0.1}, "gradeFactors": {"A": 8, "B": 4, "C": 2, "D": 0}},
        14: {"type": "linear", "terms": {"occupiedHouseholds": 0.1}},
        15: {"type": "grade-constant", "values": {"A": 18, "B": 12, "C": 8, "D": 3}},
        16: {"type": "grade-constant", "values": {"A": 6, "B": 4, "C": 4, "D": 2}},
        17: {"type": "constant", "value": 4},
        18: {"type": "constant", "value": 730},
        19: {"type": "linear", "terms": {"deliveredHouseholds": 0.05, "occupiedHouseholds": -0.05}},
        20: {"type": "linear", "terms": {"receivedHouseholds": 0.5, "occupiedHouseholds": -0.3}},
        21: {"type": "constant", "value": 12},
    }
    fixed_hours = {17: 1, 18: 2, 20: 0.1}
    rows = []
    for row in range(5, 22):
        rows.append({
            "id": f"service-{row}",
            "action": sheet.cell(row, 1).value,
            "property": sheet.cell(row, 8).value,
            "basis": sheet.cell(row, 15).value or "",
            "frequency": grade_values(sheet, row, (9, 10, 11, 12)),
            "demand": demands[row],
            "unitHours": {"type": "fixed", "value": fixed_hours[row]} if row in fixed_hours else {
                "type": "grade-adjusted",
                "base": multiplier(sheet.cell(row, 5).value),
                "additional": sheet.cell(row, 6).value or 0,
            },
        })
    return rows


CLEANING_QUANTITIES = {
    **dict.fromkeys((5, 6, 7, 12), ("perimeterEntrances", 1)),
    **dict.fromkeys((8, 9, 10), ("gateWallArea", 1)),
    11: ("gateWallArea", 0.5),
    **dict.fromkeys((13, 14, 15, 16), ("pavedRoadArea", 1)),
    **dict.fromkeys((17, 18, 19, 20, 23, 24), ("stiltFloorArea", 1)),
    **dict.fromkeys((21, 22), ("stiltWallArea", 1)),
    **dict.fromkeys((25, 26, 27, 28, 31), ("lobbyFloorArea", 1)),
    **dict.fromkeys((29, 30), ("lobbyWallArea", 1)),
    **dict.fromkeys((32, 33, 34, 37), ("standardLobbyFloorArea", 1)),
    **dict.fromkeys((35, 36), ("standardLobbyWallArea", 1)),
    **dict.fromkeys((38, 39, 40, 43), ("stairFloorArea", 1)),
    **dict.fromkeys((41, 42), ("stairWallArea", 1)),
    **dict.fromkeys((44, 45, 46), ("rooftopFloorArea", 1)),
    **dict.fromkeys((47, 48, 49, 50, 51, 52), ("garageTotalArea", 1)),
}


def cleaning_rules(sheet) -> list[dict]:
    surface = ""
    location = ""
    rows = []
    for row in range(5, 53):
        surface = sheet.cell(row, 1).value or surface
        location = sheet.cell(row, 2).value or location
        source, scale = CLEANING_QUANTITIES[row]
        rows.append({
            "id": f"cleaning-{row}",
            "action": sheet.cell(row, 3).value,
            "property": sheet.cell(row, 13).value,
            "unit": sheet.cell(row, 4).value,
            "basis": " / ".join(filter(None, (surface, location))),
            "quantitySource": source,
            "quantityScale": scale,
            "baseUnitHours": multiplier(sheet.cell(row, 6).value),
            "travelRatio": ratio(sheet.cell(row, 9).value),
            "frequency": grade_values(sheet, row, (14, 16, 18, 20)),
            "annualFrequency": grade_values(sheet, row, (15, 17, 19, 21)),
        })
    return rows


GREENING_QUANTITIES = {
    **dict.fromkeys(range(5, 13), "entranceLawnArea"),
    **dict.fromkeys((13, 14, 15, 17, 18, 19, 20, 21), "entranceGroundcoverArea"),
    16: "seasonalFlowerArea",
    22: "zero",
    **dict.fromkeys((23, 24), "entranceGreenArea"),
    **dict.fromkeys(range(25, 33), "mainLawnArea"),
    **dict.fromkeys((33, 34, 35, 37, 38, 39, 40, 41), "mainGroundcoverArea"),
    36: "zero",
    42: "zero",
    **dict.fromkeys((43, 44), "mainGreenArea"),
    **dict.fromkeys(range(45, 53), "treeShrubCount"),
    53: "zero",
    54: "zero",
    55: "zero",
}


def greening_rules(sheet) -> list[dict]:
    rows = []
    for row in range(5, 56):
        rows.append({
            "id": f"greening-{row}",
            "action": sheet.cell(row, 1).value,
            "property": sheet.cell(row, 10).value,
            "unit": sheet.cell(row, 2).value,
            "quantitySource": GREENING_QUANTITIES[row],
            "baseUnitHours": multiplier(sheet.cell(row, 4).value),
            "unitHoursScale": 0.25 if "/4" in sheet.cell(row, 9).value else 1,
            "travelRatio": ratio(sheet.cell(row, 7).value),
            "frequency": grade_values(sheet, row, (11, 13, 15, 17)),
            "annualFrequency": grade_values(sheet, row, (12, 14, 16, 18)),
        })
    return rows


def assistance_rules(sheet) -> list[dict]:
    definitions = {
        4: {"type": "multiply", "quantitySource": "one", "standards": {"A": 3, "B": 2, "C": 2, "D": 2}},
        5: {"type": "multiply", "quantitySource": "one", "standards": {"A": 0, "B": 0, "C": 0, "D": 0}},
        7: {"type": "multiply", "quantitySource": "one", "standards": {"A": 2, "B": 2, "C": 2, "D": 2}},
        8: {"type": "divide", "quantitySource": "totalBuildingArea", "standards": {"A": 80000, "B": 100000, "C": 100000, "D": 150000}},
        9: {"type": "divide", "quantitySource": "assistanceBaseRaw", "standards": {"A": 5.2, "B": 5.2, "C": 5.2, "D": 5.2}},
        10: {"type": "divide", "quantitySource": "assistanceWithReliefRaw", "standards": {"A": 4, "B": 4, "C": 4, "D": 8}},
    }
    rows = []
    for row in (4, 5, 7, 8, 9, 10):
        rule = definitions[row]
        rows.append({
            "id": f"assistance-{row}",
            "action": sheet.cell(row, 1).value,
            "property": sheet.cell(row, 4).value,
            "unit": sheet.cell(row, 2).value,
            "frequency": grade_values(sheet, row, (5, 7, 9, 11)),
            **rule,
        })
    return rows


def require_parameter_key(mapping: dict[str, str], sheet_name: str, row: int) -> str:
    source = f"{sheet_name}:{row}"
    if source not in mapping:
        raise ValueError(f"缺少高级参数映射：{source}")
    return mapping[source]


def grade_unit_hours(formula_sheet, price_sheet, row: int, column: int = 5) -> dict[str, float]:
    formula = formula_sheet.cell(row, column).value
    references = GRADE_REFERENCE.findall(formula or "")
    if len(references) != 4:
        coordinate = formula_sheet.cell(row, column).coordinate
        raise ValueError(f"无法读取四档标准工时：{formula_sheet.title}!{coordinate}")
    travel_value = formula_sheet.cell(row, column + 2).value or 0
    scale = multiplier(formula)
    base_hours = [float(price_sheet[reference].value) * scale for reference in references]
    if isinstance(travel_value, (int, float)):
        total_hours = (base + float(travel_value) for base in base_hours)
    else:
        travel_ratio = ratio(travel_value)
        total_hours = (base * (1 + travel_ratio) for base in base_hours)
    return dict(zip(GRADES, total_hours))


def property_from_action(action: str) -> str:
    if action.startswith("A-"):
        return "基础"
    if action.startswith("B-"):
        return "可选"
    raise ValueError(f"无法判定动作属性：{action}")


def engineering_rule(
    formula_sheet,
    value_sheet,
    price_sheet,
    row: int,
    prefix: str,
    parameter_key: str,
) -> dict:
    action = formula_sheet.cell(row, 1).value
    frequency = formula_sheet.cell(row, 11).value or "0"
    annual_frequency = float(value_sheet.cell(row, 14).value or 0)
    return {
        "id": f"{prefix}-{row}",
        "source": f"{formula_sheet.title}:{row}",
        "action": action,
        "system": formula_sheet.cell(row, 2).value or "",
        "property": property_from_action(action),
        "unit": formula_sheet.cell(row, 3).value or "",
        "quantityParameterKey": parameter_key,
        "templateQuantity": float(value_sheet.cell(row, 4).value or 0),
        "unitHours": grade_unit_hours(formula_sheet, price_sheet, row),
        "frequency": dict.fromkeys(GRADES, frequency),
        "annualFrequency": dict.fromkeys(GRADES, annual_frequency),
        "monthlyRate": float(value_sheet.cell(row, 16).value or 0),
    }


def pest_control_rules(formula_sheet, value_sheet, price_sheet, mapping: dict[str, str]) -> list[dict]:
    anchor_row = 5
    frequency = formula_sheet.cell(anchor_row, 10).value or "0"
    annual_frequency = float(value_sheet.cell(anchor_row, 11).value or 0)
    template_quantity = float(value_sheet.cell(anchor_row, 3).value or 0)
    unit_hours = grade_unit_hours(formula_sheet, price_sheet, anchor_row, column=4)
    rows = []
    for row in range(5, 12):
        action = formula_sheet.cell(row, 1).value
        rows.append({
            "id": f"pest-control-{row}",
            "source": f"{formula_sheet.title}:{row}",
            "action": action,
            "system": "环境消杀",
            "property": property_from_action(action),
            "unit": formula_sheet.cell(row, 2).value or "",
            "quantityParameterKey": require_parameter_key(mapping, formula_sheet.title, row),
            "templateQuantity": template_quantity,
            "unitHours": unit_hours,
            "frequency": dict.fromkeys(GRADES, frequency),
            "annualFrequency": dict.fromkeys(GRADES, annual_frequency),
            "monthlyRate": 0,
        })
    return rows


def validate_mapping(mapping: dict[str, str]) -> None:
    expected_sources = {
        f"{sheet_name}:{row}"
        for sheet_name, (start, end, _) in MISSING_SHEETS.items()
        for row in range(start, end + 1)
    }
    actual_sources = set(mapping)
    if actual_sources != expected_sources:
        missing = sorted(expected_sources - actual_sources)
        extra = sorted(actual_sources - expected_sources)
        raise ValueError(f"高级参数映射不完整；缺少={missing}；多余={extra}")
    parameter_keys = set(mapping.values())
    if len(parameter_keys) != 90:
        raise ValueError(f"高级参数数量异常：{len(parameter_keys)}")
    missing_labels = sorted(parameter_keys - PARAMETER_LABELS.keys())
    extra_labels = sorted(PARAMETER_LABELS.keys() - parameter_keys)
    if missing_labels or extra_labels:
        raise ValueError(f"高级参数名称不完整；缺少={missing_labels}；多余={extra_labels}")


def advanced_parameter_definitions(rules: list[dict], mapping: dict[str, str]) -> list[dict]:
    rules_by_key: dict[str, list[dict]] = {key: [] for key in dict.fromkeys(mapping.values())}
    for rule in rules:
        rules_by_key[rule["quantityParameterKey"]].append(rule)

    definitions = []
    for key, affected_rules in rules_by_key.items():
        units = {rule["unit"] for rule in affected_rules}
        if len(units) != 1:
            raise ValueError(f"高级参数单位不一致：{key}={sorted(units)}")
        template_values = sorted({float(rule["templateQuantity"]) for rule in affected_rules})
        if len(template_values) != 1:
            raise ValueError(f"高级参数模板数量不一致：{key}={template_values}")
        template_value = template_values[0]
        group = "grounds" if key.startswith("pest.") else key.split(".", 1)[0]
        definitions.append({
            "key": key,
            "label": PARAMETER_LABELS[key],
            "group": group,
            "unit": units.pop(),
            "templateValue": template_value,
            "templateValues": template_values,
            "defaultRule": {"type": "template", "value": template_value, "source": "template"},
            "round": "integer" if affected_rules[0]["unit"] in {"个", "台"} else "none",
            "affectedActionIds": [rule["id"] for rule in affected_rules],
        })
    return definitions


def main() -> None:
    formula_workbook = openpyxl.load_workbook(WORKBOOK, data_only=False, read_only=True)
    value_workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    mapping_document = json.loads(PARAMETER_MAP.read_text(encoding="utf-8"))
    if mapping_document.get("version") != "2026-09-full-model-v1":
        raise ValueError("高级参数映射版本不正确")
    mapping = mapping_document.get("rows", {})
    validate_mapping(mapping)

    OUTPUT.mkdir(parents=True, exist_ok=True)
    write_module("service-rules.mjs", "SERVICE_RULES", service_rules(formula_workbook["服务"]))
    write_module("cleaning-rules.mjs", "CLEANING_RULES", cleaning_rules(formula_workbook["清洁"]))
    write_module("greening-rules.mjs", "GREENING_RULES", greening_rules(formula_workbook["绿化"]))
    write_module("assistance-rules.mjs", "ASSISTANCE_RULES", assistance_rules(formula_workbook["客助"]))

    price_sheet = value_workbook[PRICE_SHEET_NAME]
    pest_rules = pest_control_rules(
        formula_workbook["四害消杀"], value_workbook["四害消杀"], price_sheet, mapping,
    )
    outsourced_rules = [
        engineering_rule(
            formula_workbook["工程委外"], value_workbook["工程委外"], price_sheet, row,
            "engineering-outsourced", require_parameter_key(mapping, "工程委外", row),
        )
        for row in range(5, 100)
    ]
    routine_rules = [
        engineering_rule(
            formula_workbook["工程常规"], value_workbook["工程常规"], price_sheet, row,
            "engineering-routine", require_parameter_key(mapping, "工程常规", row),
        )
        for row in range(5, 233)
    ]
    write_module("pest-control-rules.mjs", "PEST_CONTROL_RULES", pest_rules)
    write_module("engineering-outsourced-rules.mjs", "ENGINEERING_OUTSOURCED_RULES", outsourced_rules)
    write_module("engineering-routine-rules.mjs", "ENGINEERING_ROUTINE_RULES", routine_rules)
    definitions = advanced_parameter_definitions(
        [*pest_rules, *outsourced_rules, *routine_rules], mapping,
    )
    write_module(
        "advanced-parameter-definitions.mjs", "ADVANCED_PARAMETER_DEFINITIONS", definitions,
    )
    print(
        f"generated rules: {len(pest_rules)} / {len(outsourced_rules)} / {len(routine_rules)}; "
        f"parameters: {len(definitions)}"
    )


if __name__ == "__main__":
    main()
